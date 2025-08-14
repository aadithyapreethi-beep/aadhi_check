import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from collections import OrderedDict
import serial
import time
import os
import requests
from requests.auth import HTTPBasicAuth
import re
import logging
from datetime import datetime
import json
import subprocess

# Load configuration from JSON file
jsonFile = "config.json"
if os.path.exists(jsonFile):
    os.remove(jsonFile)
try:
    with open("config.json", "r") as f:
        config = json.load(f)
except FileNotFoundError:
    print("Error: config.json not found. Creating a default config.json")
    default_config = {
        "apiUrl": "http://172.20.97.2/rps",
        "apiUser": "root",
        "apiPass": "root",
        "baudRate": 115200,
        "iteration": 250,
        "rebootCount": 5,
        "serialDevice": "/dev/ttyUSB0",
        "username": "ubuntu",
        "password": "ubuntu123",
        "ipAddresses": ["192.168.1.11", "192.168.2.12", "192.168.3.13", "192.168.4.14", "192.168.5.15", "192.168.6.16", "192.168.7.17", "192.168.8.18"],
        "interfaces": ["eno1", "enp13s0", "enp14s0", "enp15s0", "enp16s0f0", "enp16s0f1", "enp16s0f2", "enp16s0f3"]
    }
    with open("config.json", "w") as f:
        json.dump(default_config, f, indent=4)
    config = default_config

apiUrl = config["apiUrl"]
apiUser = config["apiUser"]
apiPass = config["apiPass"]
baudRate = config["baudRate"]
iteration = config["iteration"]
rebootCount = config["rebootCount"]
serialDevice = config["serialDevice"]
username = config["username"]
password = config["password"]
ipAddresses = config["ipAddresses"]
interfaces = config["interfaces"]

# Set up logging
logFilename = "consoleLog2.txt"
logging.basicConfig(filename=logFilename, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

with open(logFilename, "w") as logFile:
    logFile.write("Starting new log...\n")


def log_message(message, level=logging.INFO):
    """Logs messages to consoleLog.txt and prints to console."""
    logging.log(level, message)
    print(message)

def clean_output(output):
    """Removes unwanted escape sequences from serial output."""
    output = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', output)
    output = re.sub(r'\x1b\[\?2004[hl]', '', output)
    return output.strip()

def read_serial_output(ser, timeout=15, command_sent=None, remove_extra_shell=True):
    """Reads and cleans serial output."""
    output = ""
    start_time = time.time()
    while time.time() - start_time < timeout:
        if ser.inWaiting() > 0:
            output += ser.read(ser.inWaiting()).decode(errors="ignore")
        time.sleep(0.1)
    output = clean_output(output)

    if command_sent:
        escaped_command = re.escape(command_sent)
        output = re.sub(rf"^\s*{escaped_command}\s*$", "", output, flags=re.MULTILINE).strip()
    return output

def extract_system_stats(output):
    """Extracts CPU Usage, Temperature, and Memory Usage from console output."""
    cpu_usage, temperature, memory_usage = "Not Found", "Not Found", "Not Found"
    output = output.replace("\r", "").replace("\n", "").replace(" ", " ")
    match = re.search(r"Usage of /:\s+([\d.]+%)", output)
    if match:
        cpu_usage = match.group(1)
    match = re.search(r"Temperature:\s+([\d.]+ C)", output)
    if match:
        temperature = match.group(1)
    match = re.search(r"Memory usage:\s+([\d.]+%)", output)
    if match:
        memory_usage = match.group(1)
    return cpu_usage, temperature, memory_usage

free_headers = [
    "before stress-ng mem total",
    "before stress-ng mem used",
    "before stress-ng mem free",
    "before stress-ng mem shared",
    "before stress-ng mem buff/cache",
    "before stress-ng mem available",
    "before stress-ng swap total",
    "before stress-ng swap used",
    "before stress-ng swap free",
    "after stress-ng mem total",
    "after stress-ng mem used",
    "after stress-ng mem free",
    "after stress-ng mem shared",
    "after stress-ng mem buff/cache",
    "after stress-ng mem available",
    "after stress-ng swap total",
    "after stress-ng swap used",
    "after stress-ng swap free",
]

def parse_free_output_hardcoded(free_output, prefix):
    """Parses the 'free -h' output with hardcoded logic."""
    free_data = {}
    lines = free_output.splitlines()

    for line in lines:
        parts = line.split()
        if len(parts) >= 7 and parts[0] == "Mem:":
            free_data[f"{prefix} mem total"] = parts[1]
            free_data[f"{prefix} mem used"] = parts[2]
            free_data[f"{prefix} mem free"] = parts[3]
            free_data[f"{prefix} mem shared"] = parts[4]
            free_data[f"{prefix} mem buff/cache"] = parts[5]
            free_data[f"{prefix} mem available"] = parts[6]
        elif len(parts) >= 4 and parts[0] == "Swap:":
            free_data[f"{prefix} swap total"] = parts[1]
            free_data[f"{prefix} swap used"] = parts[2]
            free_data[f"{prefix} swap free"] = parts[3]

    return free_data

def send_rps_command(command):
    """Sends command to RPS and handles errors."""
    try:
        response = requests.get(apiUrl, params={"set_switch": command}, auth=HTTPBasicAuth(apiUser, apiPass), timeout=10)
        response.raise_for_status()  # Raise HTTPError for bad responses (4xx or 5xx)
        log_message(f"RPS command '{command}' successful.")
        return True
    except requests.exceptions.RequestException as e:
        log_message(f"RPS command '{command}' failed. Error: {e}", level=logging.ERROR)
        return False
        
def get_random_sum_parts(total=40, count=5):
    if count > total:
        raise ValueError("Count cannot be greater than the total sum.")

    breaks = sorted(random.sample(range(1, total), count - 1))
    parts = [a - b for a, b in zip(breaks + [total], [0] + breaks)]
    return parts

# Ensure Excel file exists
excelFile = "testResults2.xlsx"
if os.path.exists(excelFile):
    os.remove(excelFile)  # Delete old file

wb = openpyxl.Workbook()
ws = wb.active

# Start with a minimal header (interfaces will be added later)
header = ["Iteration", "Timestamp", "CPU Usage", "CPU Temperature", "Memory Usage", "Power Cycle Status", "USB Status", "GPS Status"]
header_extend = ["LTE1 Status", "LTE2 Status", "SIM1 Status", "SIM2 Status", "Modem1 Connection Status",
                 "Modem2 Connection Status", "SIM1 Operator Name", "SIM2 Operator Name", "SIM1 Registration",
                 "SIM2 Registration"]
for iface in interfaces:
    header_extend.extend([f"{iface} Ethernet Detected", f"{iface} Speed", f"{iface} Ping Result"])

sensor_headers = []
free_headers = []
mpstat_headers = set()

header_full = header + header_extend + sensor_headers + free_headers + list(mpstat_headers)

ws.append(header_full)

# DYNAMIC COLUMN WIDTH ADJUSTMENT and HEADER FORMATTING
header_font = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')

# Adjust column width and apply header formatting for ALL headers
for col_num, header_item in enumerate(header_full, 1):
    column_letter = get_column_letter(col_num)
    max_length = len(header_item)  # Start with header length

    # Check data in the column for longer values
    for row in ws.iter_rows(min_row=2):
        cell_value = row[col_num - 1].value
        if cell_value and len(str(cell_value)) > max_length:
            max_length = len(str(cell_value))

    # Add some padding
    ws.column_dimensions[column_letter].width = max_length + 5

    # Apply Header formatting
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.alignment = header_alignment

# Cell alignment for data rows
data_alignment = Alignment(horizontal='center', vertical='center')

# Run for multiple iterations
for ite in range(1, iteration + 1):
    log_message(f"\n========== Iteration {ite} ==========\n")

    power_cycle_status = "Success"  # Initialize as success, will change if any reboot fails
    result = get_random_sum_parts(40, 5)
    log_message(f"\n---> result")
    # Run for multiple reboots
    for reboot_num in range(1, rebootCount + 1):
        log_message(f"\n========== Reboot {reboot_num}/{rebootCount} ==========\n")

        # Determine on and off durations for this iteration
        on_duration = reboot_num

        # Turn OFF the port
        log_message(f"Turning OFF the port (Reboot {reboot_num})...")
        if not send_rps_command("3 false"):
            power_cycle_status = "Failed"  # Update status if any reboot fails

        time.sleep(result[reboot_num])

        # Turn ON the port
        log_message(f"Turning ON the port (Reboot {reboot_num})...")
        if not send_rps_command("3 true"):
            power_cycle_status = "Failed"  # Update status if any reboot fails

        time.sleep(on_duration)

    log_message("Reboot cycle completed.")

    log_message("Waiting for 240 seconds for device reboot...")
    time.sleep(240)

    # Establish serial connection
    log_message("Re-establishing serial connection...")
    try:
        ser = serial.Serial(serialDevice, baudRate, timeout=1)
        time.sleep(2)
    except serial.SerialException as e:
        log_message(f"Error opening serial port: {e}", level=logging.ERROR)
        continue

    # Ensure login prompt
    log_message("Checking for 'ubuntu login:' prompt...")
    ser.write(b"\n" * 5)
    output = read_serial_output(ser, 30, remove_extra_shell=False)
    if "ubuntu login:" not in output:
        log_message("[ERROR] 'ubuntu login:' not detected, skipping this iteration.", level=logging.ERROR)
        continue

    # Enter credentials
    log_message("Logging in...")
    ser.write(f"{username}\n".encode())
    output = read_serial_output(ser, 10, remove_extra_shell=False)
    if "Password:" not in output:
        log_message("[ERROR] Password prompt not detected, skipping this iteration.", level=logging.ERROR)
        continue
    ser.write(f"{password}\n".encode())
    output = read_serial_output(ser, 10)

    # Check for login prompt
    if "root@ubuntu:" not in output and "ubuntu@ubuntu:" not in output:
        log_message("[ERROR] Login failed, skipping this iteration.", level=logging.ERROR)
        continue

    # Extract system stats
    cpu_usage, temperature, memory_usage = extract_system_stats(output)
    log_message(f"CPU Usage: {cpu_usage}, CPU Temperature: {temperature}, Memory Usage: {memory_usage}")

    ser.write("sudo su\n".encode())
    time.sleep(2)
    ser.write(f"{password}\n".encode())
    time.sleep(2)

    # Clear the sensor data before each iteration.
    sensors_data = {}
    free_data = {}
    mpstat_data = {}

    sensor_headers = []
    free_headers = []
    mpstat_headers = OrderedDict()

    # Sensor data collection (before stress-ng)
    log_message("Collecting sensor data (before stress-ng)...")
    ser.write(b"sensors\n")
    sensors_output = read_serial_output(ser, 10)
    log_message(sensors_output)

    lines = sensors_output.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if ":" in line and "Adapter:" not in line and "[sudo]" not in line and "root@ubuntu" not in line:
            parts = line.split(":")
            sensor_name = parts[0].strip()
            sensor_value = parts[1].split("(")[0].strip()

            full_sensor_name = sensor_name
            adapter_info = None

            # Scan backward to find adapter information
            for j in range(i - 1, -1, -1):
                if "Adapter:" in lines[j]:
                    adapter_info = lines[j - 1].strip() + " " + lines[j].split(":")[1].strip()
                    break

            if adapter_info:
                full_sensor_name = "before stress-ng " + adapter_info + " " + sensor_name
            elif "Core" in sensor_name:
                full_sensor_name = "before stress-ng " + sensor_name
            else:
                full_sensor_name = "before stress-ng " + sensor_name

            # Check for existing detailed header before adding simpler one
            if adapter_info or "Core" in sensor_name:
                sensors_data[full_sensor_name] = sensor_value
            elif full_sensor_name not in [key for key in sensors_data.keys() if "Adapter:" in key or "Core" in key]:
                sensors_data[full_sensor_name] = sensor_value

        i += 1

    # Memory usage details (free -h) (before stress-ng)
    log_message("Collecting memory usage details (free -h) (before stress-ng)...")
    command = f"echo {password} | sudo -S free -h"
    ser.write(f"{command}\n".encode())
    free_output = read_serial_output(ser, 10, command_sent=command)
    log_message(free_output)

    free_data.update(parse_free_output_hardcoded(free_output, "before stress-ng"))

    # mpstat -P ALL 1 1 (before stress-ng)
    log_message("Collecting mpstat -P ALL 1 1 data (before stress-ng)...")
    ser.write(b"mpstat -P ALL 1 1\n")
    mpstat_output = read_serial_output(ser, 10)
    log_message(mpstat_output)

    mpstat_lines = mpstat_output.splitlines()
    for line in mpstat_lines:
        if line.startswith("Average:") or line.startswith("1"):
            parts = line.split()
            if len(parts) >= 12 and parts[1] != "CPU":
                cpu_name = parts[1]
                mpstat_details = {
                    f"before stress-ng average CPU {cpu_name} %usr": parts[2],
                    f"before stress-ng average CPU {cpu_name} %nice": parts[3],
                    f"before stress-ng average CPU {cpu_name} %sys": parts[4],
                    f"before stress-ng average CPU {cpu_name} %iowait": parts[5],
                    f"before stress-ng average CPU {cpu_name} %soft": parts[7],
                    f"before stress-ng average CPU {cpu_name} %idle": parts[11],
                }
                mpstat_data.update(mpstat_details)
                mpstat_headers.update(mpstat_details)

    # stress-ng commands
    log_message("Starting stress-ng commands...")
    ser.write(b"nohup stress-ng --vm $(nproc) --vm-bytes 100% --timeout 14m &\n")
    time.sleep(2)  # give time for command to start.
    ser.write(b"nohup stress-ng --cpu $(nproc) --timeout 14m &\n")
    time.sleep(5)  # give time for command to start.
    log_message("stress-ng commands started.")

    # Sensor data collection (after stress-ng)
    log_message("Collecting sensor data (after stress-ng)...")
    ser.write(b"sensors\n")
    sensors_output = read_serial_output(ser, 10)
    log_message(sensors_output)

    lines = sensors_output.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if ":" in line and "Adapter:" not in line and "[sudo]" not in line and "root@ubuntu" not in line:
            parts = line.split(":")
            sensor_name = parts[0].strip()
            sensor_value = parts[1].split("(")[0].strip()

            full_sensor_name = sensor_name
            adapter_info = None

            # Scan backward to find adapter information
            for j in range(i - 1, -1, -1):
                if "Adapter:" in lines[j]:
                    adapter_info = lines[j - 1].strip() + " " + lines[j].split(":")[1].strip()
                    break

            if adapter_info:
                full_sensor_name = "after stress-ng " + adapter_info + " " + sensor_name
            elif "Core" in sensor_name:
                full_sensor_name = "after stress-ng " + sensor_name
            else:
                full_sensor_name = "after stress-ng " + sensor_name

            # Check for existing detailed header before adding simpler one
            if adapter_info or "Core" in sensor_name:
                sensors_data[full_sensor_name] = sensor_value
            elif full_sensor_name not in [key for key in sensors_data.keys() if "Adapter:" in key or "Core" in key]:
                sensors_data[full_sensor_name] = sensor_value

        i += 1

    # Memory usage details (free -h) (after stress-ng)
    log_message("Collecting memory usage details (free -h) (after stress-ng)...")
    command = f"echo {password} | sudo -S free -h"
    ser.write(f"{command}\n".encode())
    free_output = read_serial_output(ser, 10, command_sent=command)
    log_message(free_output)

    free_data.update(parse_free_output_hardcoded(free_output, "after stress-ng"))

    # mpstat -P ALL 1 1 (after stress-ng)
    log_message("Collecting mpstat -P ALL 1 1 data (after stress-ng)...")
    ser.write(b"mpstat -P ALL 1 1\n")
    mpstat_output = read_serial_output(ser, 10)
    log_message(mpstat_output)

    mpstat_lines = mpstat_output.splitlines()
    for line in mpstat_lines:
        if line.startswith("Average:") or line.startswith("1"):
            parts = line.split()
            if len(parts) >= 12 and parts[1] != "CPU":
                cpu_name = parts[1]
                mpstat_details = {
                    f"after stress-ng average CPU {cpu_name} %usr": parts[2],
                    f"after stress-ng average CPU {cpu_name} %nice": parts[3],
                    f"after stress-ng average CPU {cpu_name} %sys": parts[4],
                    f"after stress-ng average CPU {cpu_name} %iowait": parts[5],
                    f"after stress-ng average CPU {cpu_name} %soft": parts[7],
                    f"after stress-ng average CPU {cpu_name} %idle": parts[11],
                }
                mpstat_data.update(mpstat_details)
                mpstat_headers.update(mpstat_details)

    # Update sensor headers and data
    sensor_headers = list(sensors_data.keys())
    free_headers = list(free_data.keys())
    mpstat_headers = list(mpstat_headers.keys()) #convert back to list to keep it consistent
    mpstat_headers.sort()

    header_full = header + header_extend + sensor_headers + free_headers + mpstat_headers

    # Update header in excel
    for col_num, header_item in enumerate(header_full, 1):
        ws.cell(row=1, column=col_num, value=header_item)

    # After the headers are added, re-apply formatting and column width adjustments.
    for col_num, header_item in enumerate(header_full, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(header_item)

        for row in ws.iter_rows(min_row=2):
            cell_value = row[col_num - 1].value
            if cell_value and len(str(cell_value)) > max_length:
                max_length = len(str(cell_value))

        ws.column_dimensions[column_letter].width = max_length + 5

        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment

    # eMMC command
    log_message("Starting eMMC command...")
    ser.write(b"./eMMC_aggressive_3mins.sh &\n")
    time.sleep(60)  # give time for command to start.

    # Configure network interfaces
    log_message("Configuring Network Interfaces...")

    # Configure network interfaces with retry mechanism
    interface_results = {}
    for i, iface in enumerate(interfaces):
        log_message(f"Attempting to bring up interface {iface}...")
        for retry in range(3):
            ser.write(f"echo {password} | sudo -S ip link set {iface} up\n".encode())
            time.sleep(5)  # Increased sleep time

            if i < len(ipAddresses):
                log_message(f"Assigning IP {ipAddresses[i]} to {iface}...")
                ser.write(f"echo {password} | sudo -S ip addr add {ipAddresses[i]}/24 dev {iface}\n".encode())
                time.sleep(5)  # Increased sleep time

            log_message(f"Checking operational state of {iface} (attempt {retry + 1})...")
            ser.write(f"echo {password} | sudo -S ethtool {iface}\n".encode())
            time.sleep(5)  # Increased sleep time

            ethtool_output = read_serial_output(ser, 5)

            link_detected_match = re.search(r"Link detected: yes", ethtool_output)
            speed_match = re.search(r"Speed:\s*(\d+)Mb/s", ethtool_output)

            link_status = "Yes" if link_detected_match else "No"
            speed = speed_match.group(0) if speed_match else "Unknown"

            log_message(f"Interface {iface} - Link Detected: {link_status}, Speed: {speed}")

            if link_detected_match:
                if retry > 0:
                    log_message(f"Interface {iface} came up after {retry+1} retries in iteration {ite}")
                break  # Interface came up, break retry loop
            elif retry == 2:  # All retries failed
                log_message(f"Interface {iface} failed to come up after 3 retries in iteration {ite}")
                break

        if link_detected_match and i < len(ipAddresses):
            log_message(f"Pinging {ipAddresses[i]} from {iface}...")
            ser.write(f"ping -c 10 {ipAddresses[i]}\n".encode())
            time.sleep(15)

            ping_output = read_serial_output(ser, 15)
            log_message(ping_output)

            match = re.search(r"(\d+) packets transmitted, (\d+) received, (\d+)% packet loss", ping_output)

            if match:
                _, packets_received, packet_loss = map(int, match.groups())
                ping_result = "Ping Passed" if packet_loss == 0 else "Ping Failed"
            else:
                ping_result = "Ping Failed"

            log_message(f"Ping Result for {iface}: {ping_result}")
        elif link_status == "No":
            ping_result = "Link Down"
        else:
            ping_result = "Speed Not 1000Mb/s"

        # Update interface_results for excel sheet.
        if iface not in interface_results:
            interface_results[iface] = {}
        interface_results[iface]["Ethernet Detected"] = link_status
        interface_results[iface]["Speed"] = speed
        interface_results[iface]["Ping Result"] = ping_result

    # Step 7: USB Check
    log_message("Checking for USB device (/dev/sda)...")
    ser.write(b"ls /dev/sd*\n")
    usb_output = read_serial_output(ser, 5)
    log_message(usb_output)
    usb_status = "USB Found" if "/dev/sda" in usb_output else "USB Not Found"
    log_message(f"USB Status: {usb_status}")

    # Step 10: Check the interface was created for the modules (for both LTEs)
    lte_interfaces = ["LTE Interface Not Found", "LTE Interface Not Found"]  # prefill with default values.
    lte_devices_count = 0  # count the Telit devices.
    for lte_num in range(2):  # Assuming 2 LTEs, adjust as needed
        # Step 8: LTE (Long Term Evolution) Check
        log_message("Checking for LTE")
        ser.write(b"lsusb\n")
        lte_output = read_serial_output(ser, 5)
        log_message(lte_output)
        lte_status = "LTE Found" if "Telit Wireless Solutions" in lte_output else "LTE Not Found"
        if "Telit Wireless Solutions" in lte_output:
            lte_devices_count += 1
        log_message(f"LTE Status: {lte_status}")
        log_message(f"Checking LTE interface {lte_num + 1} was created or not")
        ser.write(b"ip link show\n")
        interface_output = read_serial_output(ser, 5)
        log_message(interface_output)

        log_message(f"LTE Interface {lte_num + 1}: {lte_interfaces[lte_num]}")  # display the default value.

    # Step 12: Checking modem status (for both modems)
    modem_connection_statuses = []
    for modem_num in range(2):  # Assuming 2 modems, adjust as needed
        log_message(f"Checking Modem {modem_num + 1} Status...")

        # Determine the correct modem index (always 0 or 1)
        modem_index = str(modem_num)

        # Attempt to connect the modem with retry
        connection_status = "Connection Failed"
        for retry in range(3):
            ser.write(f"mmcli -m {modem_index}\n".encode())
            modem_output = read_serial_output(ser, 5)

            connect_command = f"sudo mmcli -m {modem_index} --simple-connect=\"apn=airtelgprs.com\"\n"
            ser.write(connect_command.encode())
            modem_status_output = read_serial_output(ser, 5)
            if "password for ubuntu" in modem_status_output.lower():
                log_message("Sudo password required, sending password...")
                ser.write(f"{password}\n".encode())  # Sending the password
                modem_status_output = read_serial_output(ser, 5)  # Read the output again

            log_message(modem_status_output)

            # Check for successful connection
            if "successfully connected the modem" in modem_status_output.lower():
                connection_status = "Connected Successfully"
                break
            time.sleep(3)

        log_message(f"Modem {modem_num + 1} Connection Status: {connection_status}")
        modem_connection_statuses.append(connection_status)

    # Step 11: Identify the USB Ports Mapped for AT Command Communication and Verify the Module Loaded for the SIM (for both SIMs)
    sim_statuses = []
    for sim_num in range(2):  # Assuming 2 SIMs, adjust as needed
        log_message(f"Checking SIM {sim_num + 1} Status...")

        # Send mmcli command (always -m 0 or -m 1)
        ser.write(f"mmcli -m {sim_num}\n".encode())
        sim_output = read_serial_output(ser, 5)

        log_message(sim_output)

        # Default values
        sim_status = "SIM Not Detected"
        sim_slot = "Unknown"

        # Extract SIM Slot & Status
        match = re.search(r"slot (\d+):\s*(.+?)\s*\(active\)", sim_output)
        if match:
            sim_slot = f"Slot {match.group(1)}"
            if "none" in match.group(2).lower():
                sim_status = "SIM Not Detected"
            else:
                sim_status = "SIM Active"

        print(f"SIM Slot: {sim_slot}, SIM Status: {sim_status}")
        sim_statuses.append({"Slot": sim_slot, "Status": sim_status})

    # Step 13: Get the sim details (for both SIMs)
    sim_details = []
    for modem_num in range(2):  # Assuming 2 modems, adjust as needed
        log_message(f"Getting SIM {modem_num + 1} Details...")

        # Determine the correct modem index (always 0 or 1)
        modem_index = str(modem_num)
        ser.write(f"mmcli -m {modem_index}\n".encode())
        modem_output = read_serial_output(ser, 5)

        # Run command to get SIM details
        get_sim_command = f"sudo mmcli -m {modem_index}\n"
        ser.write(get_sim_command.encode())

        # Check if sudo password is required
        sim_details_output = read_serial_output(ser, 5)
        if "password for ubuntu" in sim_details_output.lower():
            log_message("Sudo password required, sending password...")
            ser.write(f"{password}\n".encode())
            sim_details_output = read_serial_output(ser, 5)

        log_message(sim_details_output)

        # Extract the required details
        operator_name = f"SIM{modem_num + 1} Operator Name Not Found"
        registration_status = f"SIM{modem_num + 1} Registration Not Found"

        details = {"Operator Name": operator_name, "Registration": registration_status}
        if "operator name" in sim_details_output.lower() and "registration" in sim_details_output.lower():
            try:
                for line in sim_details_output.split("\n"):
                    if "operator name" in line:
                        operator_name = line.split(":")[-1].strip()
                        details["Operator Name"] = operator_name
                    elif "registration" in line:
                        registration_status = line.split(":")[-1].strip()
                        details["Registration"] = registration_status
            except Exception as e:
                log_message(f"Error parsing SIM {modem_num + 1} details: {e}", level=logging.ERROR)
        else:
            log_message(f"SIM {modem_num + 1} operator name or registration not found")

        # Log the extracted values
        log_message(f"SIM {modem_num + 1} Operator Name: {details['Operator Name']}, Registration: {details['Registration']}")
        sim_details.append(details)

    # Step GPS Check
    log_message("Checking GPS Status...")
    ser.write(b"lsusb\n")
    gps_output = read_serial_output(ser, 5)
    gps_status = "GPS Found" if "U-Blox AG u-blox GNSS receiver" in gps_output else "GPS Not Found"
    log_message(f"GPS Status: {gps_status}")

    # Step 14: Store data in Excel
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_data = [ite, timestamp, cpu_usage, temperature, memory_usage, power_cycle_status, usb_status, gps_status]

    # Add LTE and SIM results to row_data
    if lte_devices_count == 1 and len(lte_interfaces) == 1:
        lte1_status = "LTE Found"
        lte2_status = "LTE Not Found"
    elif lte_devices_count == 2 and len(lte_interfaces) == 2:
        lte1_status = "LTE Found"
        lte2_status = "LTE Found"
    else:
        lte1_status = "LTE Not Found"
        lte2_status = "LTE Not Found"

    row_data.extend([lte1_status, lte2_status, sim_statuses[0]["Status"], sim_statuses[1]["Status"],
                     modem_connection_statuses[0], modem_connection_statuses[1],
                     sim_details[0]["Operator Name"], sim_details[1]["Operator Name"],
                     sim_details[0]["Registration"], sim_details[1]["Registration"]])

    # Add interface results to row_data
    for iface in interfaces:
        row_data.extend([interface_results[iface]["Ethernet Detected"],
                         interface_results[iface]["Speed"],
                         interface_results[iface]["Ping Result"]])

    sensor_headers = list(sensors_data.keys())
    free_headers = list(free_data.keys())
    mpstat_headers = list(mpstat_headers)

    header_full = header + header_extend + sensor_headers + free_headers + mpstat_headers

    # Update header in excel
    for col_num, header_item in enumerate(header_full, 1):
        ws.cell(row=1, column=col_num, value=header_item)

    # After the headers are added, re-apply formatting and column width adjustments.
    for col_num, header_item in enumerate(header_full, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(header_item)

        for row in ws.iter_rows(min_row=2):
            cell_value = row[col_num - 1].value
            if cell_value and len(str(cell_value)) > max_length:
                max_length = len(str(cell_value))

        ws.column_dimensions[column_letter].width = max_length + 5

        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment

    # Add sensor, free, and mpstat data to row_data
    for sensor in sensor_headers:
        row_data.append(sensors_data[sensor])
    for free in free_headers:
        row_data.append(free_data[free])
    for mpstat in mpstat_headers:
        row_data.append(mpstat_data[mpstat])

    log_message(f"row_data: {row_data}") # add this line.

    header_full = header + header_extend + sensor_headers + free_headers + mpstat_headers
    log_message(f"header_full: {header_full}") # add this line.

    ws.append(row_data)

    # Apply data alignment for all cells in the row
    for col_num in range(1, len(row_data) + 1):
        cell = ws.cell(row=ite + 1, column=col_num)
        cell.alignment = data_alignment
        
    ser.write("rm /mnt/lvm/testfile\n".encode())
    time.sleep(2)

    # ser.write("exit\n".encode())
    # time.sleep(1)
    # ser.write("exit\n".encode())
    # time.sleep(1)
    # ser.write("exit\n".encode())
    # time.sleep(1)
    # ser.write(b"\n" * 5)
    # time.sleep(1)

    # Save the Excel file after each iteration
    try:
        wb.save(excelFile)
        log_message(f"Results saved to {excelFile} after iteration {ite}")
    except Exception as e:
        log_message(f"Error saving Excel file after iteration {ite}: {e}", level=logging.ERROR)

    ser.close()

# Save the Excel file
#wb.save(excelFile)
#log_message(f"Results saved to {excelFile}")
